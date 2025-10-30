import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import os

# Tên file Excel của bạn
EXCEL_FILE_SOURCE = 'CodeReviews.xlsx'
EXCEL_FILE_DEST = 'Code_Review_Report_Horizontal.xlsx' # Đổi tên file output
OUTPUT_SHEET_NAME = 'Summary_Report'

# ... (Phần 1, 2, 3: Đọc và tính toán dữ liệu không đổi - Giả sử bạn đã dán toàn bộ logic này) ...

# 1. Đọc tất cả các sheet
try:
    all_sheets = pd.read_excel(EXCEL_FILE_SOURCE, sheet_name=None, header=None, engine='openpyxl')
except FileNotFoundError:
    print(f"Lỗi: Không tìm thấy file nguồn '{EXCEL_FILE_SOURCE}'.")
    exit()

# Lọc sheet báo cáo (tên bắt đầu bằng 'Ass1')
report_sheets = {name: df for name, df in all_sheets.items() if name.startswith('Ass1')}

# Lấy Check List
try:
    check_list_df = all_sheets['Check list'].iloc[0:, 0:2].dropna(subset=[0]).rename(columns={0: 'Check Code', 1: 'Description'})
    check_list = check_list_df[pd.to_numeric(check_list_df['Check Code'], errors='coerce').notna()].copy()
    check_list['Check Code'] = check_list['Check Code'].astype('Int64') 
    check_list = check_list.set_index('Check Code')['Description'].to_dict()
except KeyError:
    print("Lỗi: Không tìm thấy sheet 'Check list'.")
    exit()

# Khởi tạo các danh sách tổng hợp
file_errors = []
top_errors_map = {}
reviewer_stats = {}

# 2. Xử lý từng file báo cáo (sheet)
for file_name, df_full in report_sheets.items():
    try:
        df_report = df_full.iloc[11:].iloc[:, [0, 5]].dropna(subset=[0, 5]) 
        df_report.columns = ['Check Code', 'Reviewer']
    except IndexError:
        continue

    df_report['Reviewer'] = df_report['Reviewer'].astype(str).str.strip()
    df_report['Check Code'] = pd.to_numeric(df_report['Check Code'], errors='coerce').astype('Int64')
    df_report = df_report.dropna(subset=['Check Code', 'Reviewer'])
    
    if df_report.empty:
        continue

    unique_codes = df_report['Check Code'].unique()
    file_errors.append({'File/Sheet Name': file_name, 'Unique Errors Count': len(unique_codes)})
    for code in unique_codes:
        code_int = int(code)
        top_errors_map[code_int] = top_errors_map.get(code_int, 0) + 1
        
    df_unique_per_reviewer_per_file = df_report.groupby(['Reviewer', 'Check Code']).size().reset_index(name='Count')
    file_reviewers = df_report['Reviewer'].unique()

    for index, row in df_unique_per_reviewer_per_file.iterrows():
        reviewer = row['Reviewer']
        if reviewer not in reviewer_stats:
            reviewer_stats[reviewer] = {'Unique Files Reviewed': 0, 'Total Unique Errors Reported (Per File)': 0}
        reviewer_stats[reviewer]['Total Unique Errors Reported (Per File)'] += 1
        
    for reviewer in file_reviewers:
        if reviewer not in reviewer_stats:
            reviewer_stats[reviewer] = {'Unique Files Reviewed': 0, 'Total Unique Errors Reported (Per File)': 0}
        reviewer_stats[reviewer]['Unique Files Reviewed'] += 1

# 3. Tạo DataFrames cho các bảng thống kê cuối cùng
df_file_stats = pd.DataFrame(file_errors)
df_top_errors = pd.DataFrame(list(top_errors_map.items()), columns=['Check Code', 'Total Files Affected'])
df_top_errors.insert(1, 'Description', df_top_errors['Check Code'].map(check_list).fillna('Description Missing'))
df_top_errors = df_top_errors.sort_values(by='Total Files Affected', ascending=False)
df_reviewer_stats = pd.DataFrame.from_dict(reviewer_stats, orient='index').reset_index().rename(columns={'index': 'Reviewer'})
df_reviewer_stats = df_reviewer_stats.sort_values(by='Total Unique Errors Reported (Per File)', ascending=False)


# 4. Sao chép và Ghi kết quả vào file Excel MỚI (CHỈNH SỬA LỚN Ở ĐÂY)
if not os.path.exists(EXCEL_FILE_SOURCE):
    print("Lỗi: Không tìm thấy file nguồn để sao chép.")
    exit()

try:
    shutil.copyfile(EXCEL_FILE_SOURCE, EXCEL_FILE_DEST)
except Exception as e:
    print(f"Lỗi khi sao chép file: {e}")
    exit()

# Mở file mới bằng Openpyxl
try:
    wb = load_workbook(EXCEL_FILE_DEST)
    
    # Xóa sheet Summary cũ nếu tồn tại
    if OUTPUT_SHEET_NAME in wb.sheetnames:
        del wb[OUTPUT_SHEET_NAME]
    
    # Tạo sheet Summary mới và đặt ở vị trí đầu tiên
    ws = wb.create_sheet(OUTPUT_SHEET_NAME, 0) 

    # Khởi tạo vị trí bắt đầu (Hàng 1, Cột 1)
    START_ROW = 1 
    current_col_idx = 1 
    
    # Hàm ghi tiêu đề
    def write_title(title, sheet, row_idx, col_idx):
        sheet.cell(row=row_idx, column=col_idx, value=title)
        return row_idx # Trả về hàng không đổi

    # Hàm ghi DataFrame
    def write_dataframe(df, sheet, row_idx, col_idx):
        # Ghi headers (tiêu đề cột của DataFrame)
        for c_idx, col_name in enumerate(df.columns):
            sheet.cell(row=row_idx, column=col_idx + c_idx, value=col_name)
            
        # Ghi dữ liệu
        for r_idx, row in enumerate(dataframe_to_rows(df, header=False, index=False)):
            for c_idx, value in enumerate(row):
                 sheet.cell(row=row_idx + r_idx + 1, column=col_idx + c_idx, value=value)
                 
        # Trả về cột tiếp theo sau khi ghi (+1 cột trống để tách bảng)
        return col_idx + len(df.columns) + 1 

    
    # --- 1. Top Lỗi ---
    if not df_top_errors.empty:
        write_title('Bảng 1: Top Lỗi (Theo số lượng File bị ảnh hưởng)', ws, START_ROW, current_col_idx)
        current_col_idx = write_dataframe(df_top_errors, ws, START_ROW + 1, current_col_idx) # +1 hàng cho tiêu đề bảng
    
    # --- 2. Số lỗi theo File ---
    if not df_file_stats.empty:
        write_title('Bảng 2: Số lỗi DUY NHẤT theo File', ws, START_ROW, current_col_idx)
        current_col_idx = write_dataframe(df_file_stats, ws, START_ROW + 1, current_col_idx)
        
    # --- 3 & 4. Thống kê theo Reviewer ---
    if not df_reviewer_stats.empty:
        write_title('Bảng 3 & 4: Thống kê theo Reviewer (Lỗi duy nhất trong mỗi File)', ws, START_ROW, current_col_idx)
        current_col_idx = write_dataframe(df_reviewer_stats, ws, START_ROW + 1, current_col_idx)

    # Lưu file
    wb.save(EXCEL_FILE_DEST)

except Exception as e:
    print(f"Lỗi khi ghi dữ liệu bằng Openpyxl: {e}")
    
print(f"\n✅ Đã hoàn thành. File '{EXCEL_FILE_DEST}' đã được tạo.")
print(f"Các bảng thống kê được sắp xếp theo **chiều ngang** trên Sheet '{OUTPUT_SHEET_NAME}'.")